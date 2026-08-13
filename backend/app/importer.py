from __future__ import annotations

import argparse
import csv
import json
import re
from io import BytesIO
from openpyxl import Workbook, load_workbook
from pathlib import Path

from sqlalchemy import select

from bs4 import BeautifulSoup

from app.catalog import map_field, seed_spec_template, upsert_model
from app.db import SessionLocal
from app.models import CpuCompatibility
from app.schemas import ModelWrite, SpecInput, ImportPreviewRow, ImportPreviewSpecRow
from app.security import ensure_default_api_client

MISSING = "待补充"


IMPORT_TEMPLATE_HEADERS = [
    "品牌代码", "品牌名称", "产品类型", "系列", "型号", "标题", "平台厂商", "代际",
    "官网参数链接", "产品彩页链接", "产品技术白皮书下载", "来源引用", "原始来源ID",
]

IMPORT_SPEC_HEADERS = ["型号", "字段分组", "字段标签", "field_key", "值", "来源引用"]
REQUIRED_MODEL_HEADERS = ["品牌代码", "品牌名称", "产品类型", "系列", "型号"]
REQUIRED_SPEC_HEADERS = ["型号", "字段分组", "字段标签", "field_key", "值", "来源引用"]


BRAND_CODE_ALIASES = {
    "len": "lenovo",
    "lk": "lenovo",
    "联想": "lenovo",
    "联想开天": "lenovo",
    "inspur": "inspur",
    "浪潮": "inspur",
    "dell": "dell",
    "戴尔": "dell",
    "generic": "generic",
    "示例品牌": "generic",
    "accessory": "accessory",
    "配件": "accessory",
}


def normalize_import_brand(code: str, name: str | None = None) -> tuple[str, str | None]:
    raw = (code or "").strip()
    key = raw.lower() if raw else (name or "").strip().lower()
    normalized = BRAND_CODE_ALIASES.get(key, raw)
    if normalized == "lenovo" and (not name or name.upper() in {"LEN", "LK"}):
        name = "联想"
    if normalized == "generic" and not name:
        name = "示例品牌"
    return normalized, name


def cell_text(values: list, idx: dict[str, int], name: str, default: str = "") -> str:
    pos = idx.get(name)
    if pos is None or pos < 0 or pos >= len(values):
        return default
    value = values[pos]
    return str(value).strip() if value is not None else default


def build_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "型号主数据"
    ws.append(IMPORT_TEMPLATE_HEADERS)
    ws.append(["inspur", "浪潮", "服务器", "NF5280", "NF5280M7", "示例标题", "Intel", "G7", "https://example.com/params", "https://example.com/brochure.pdf", "https://example.com/whitepaper.pdf", "manual", "SRC-001"])
    ws2 = wb.create_sheet("规格明细")
    ws2.append(IMPORT_SPEC_HEADERS)
    ws2.append(["NF5280M7", "基础信息", "选型注意事项", "selection_notes", "可结合项目场景填写。", "manual"])
    ws2.append(["NF5280M7", "基础信息", "官网参数链接", "official_params_url", "https://example.com/params", "manual"])
    ws2.append(["NF5280M7", "基础信息", "产品彩页", "product_brochure_url", "https://example.com/brochure.pdf", "manual"])
    ws2.append(["NF5280M7", "基础信息", "产品技术白皮书下载", "whitepaper_url", "https://example.com/whitepaper.pdf", "manual"])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def parse_import_workbook(data: bytes):
    errors: list[str] = []
    rows: list[ModelWrite] = []
    sheet_rows: list[ImportPreviewSpecRow] = []
    try:
        wb = load_workbook(filename=BytesIO(data), data_only=True)
    except Exception as exc:
        return [], [f"无法读取 Excel 文件：{exc}"], []

    if "型号主数据" not in wb.sheetnames:
        return [], ["缺少“型号主数据”工作表"], []
    if "规格明细" not in wb.sheetnames:
        errors.append("缺少“规格明细”工作表")

    ws = wb["型号主数据"]
    header = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header) if name}
    for name in REQUIRED_MODEL_HEADERS:
        if name not in idx:
            errors.append(f"型号主数据缺少必要列：{name}")
    if any(name not in idx for name in REQUIRED_MODEL_HEADERS):
        return [], errors, []

    for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in row]
        if not any(str(v or "").strip() for v in values):
            continue
        missing = [name for name in REQUIRED_MODEL_HEADERS if not cell_text(values, idx, name)]
        if missing:
            errors.append(f"型号主数据第 {row_number} 行缺少必要字段：{', '.join(missing)}")
            continue
        model_name = cell_text(values, idx, "型号")
        brand_code, brand_name = normalize_import_brand(cell_text(values, idx, "品牌代码"), cell_text(values, idx, "品牌名称") or None)
        rows.append(ModelWrite(
            brand_code=brand_code,
            brand_name=brand_name,
            product_type=cell_text(values, idx, "产品类型"),
            series=cell_text(values, idx, "系列"),
            model_name=model_name,
            title=cell_text(values, idx, "标题") or model_name,
            platform_vendor=cell_text(values, idx, "平台厂商") or None,
            generation=cell_text(values, idx, "代际") or None,
            source_ref=cell_text(values, idx, "来源引用") or "excel",
            raw_source_id=cell_text(values, idx, "原始来源ID") or None,
            specifications=[],
        ))

    if "规格明细" in wb.sheetnames:
        ws2 = wb["规格明细"]
        header2 = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws2.iter_rows(min_row=1, max_row=1))]
        idx2 = {name: i for i, name in enumerate(header2) if name}
        missing_headers = [name for name in REQUIRED_SPEC_HEADERS if name not in idx2]
        if missing_headers:
            errors.append(f"规格明细工作表表头不完整，缺少：{', '.join(missing_headers)}")
        else:
            for row_number, row in enumerate(ws2.iter_rows(min_row=2), start=2):
                vals = [cell.value for cell in row]
                if not any(str(v or "").strip() for v in vals):
                    continue
                missing = [name for name in ["型号", "字段标签", "field_key", "值"] if not cell_text(vals, idx2, name)]
                if missing:
                    errors.append(f"规格明细第 {row_number} 行缺少必要字段：{', '.join(missing)}")
                    continue
                sheet_rows.append(ImportPreviewSpecRow(
                    row_number=row_number,
                    model_name=cell_text(vals, idx2, "型号"),
                    field_group=cell_text(vals, idx2, "字段分组") or "其他",
                    field_label=cell_text(vals, idx2, "字段标签"),
                    field_key=cell_text(vals, idx2, "field_key"),
                    value=cell_text(vals, idx2, "值"),
                    source_ref=cell_text(vals, idx2, "来源引用") or "excel",
                ))
    return rows, errors, sheet_rows


GROUP_NAME_BY_CODE = {
    "basic": "基础信息",
    "processor": "处理器",
    "memory": "内存",
    "storage": "存储",
    "raid": "RAID",
    "network": "网络",
    "pcie_expansion": "PCIe与扩展",
    "gpu": "GPU",
    "power": "电源",
    "management": "管理",
    "dimension_environment": "尺寸与环境",
    "os_certification": "操作系统与认证",
    "other": "其他",
}

MODEL_META_ALIASES = {
    "品牌代码": "brand_code",
    "brand_code": "brand_code",
    "品牌": "brand_name",
    "品牌名称": "brand_name",
    "brand_name": "brand_name",
    "产品类型": "product_type",
    "product_type": "product_type",
    "系列": "series",
    "series": "series",
    "型号": "model_name",
    "model": "model_name",
    "model_name": "model_name",
    "标题": "title",
    "title": "title",
    "平台厂商": "platform_vendor",
    "platform_vendor": "platform_vendor",
    "代际": "generation",
    "generation": "generation",
    "来源引用": "source_ref",
    "source_ref": "source_ref",
    "原始来源ID": "raw_source_id",
    "raw_source_id": "raw_source_id",
}


def _strip_markdown_bullet(line: str) -> str:
    return re.sub(r"^\s*(?:[-*+]\s+|\d+[.)、]\s*)", "", line or "").strip()


def _split_markdown_key_value(line: str) -> tuple[str, str] | None:
    line = _strip_markdown_bullet(line)
    if not line or line.startswith("|"):
        return None
    match = re.match(r"^\s*([^：:\t|]{1,80})\s*(?:[:：\t|]|\s{2,})\s*(.+?)\s*$", line)
    if not match:
        return None
    return match.group(1).strip(), match.group(2).strip()


def _markdown_section_title(line: str) -> str | None:
    m = re.match(r"^\s*#{2,6}\s+(.+?)\s*$", line or "")
    return m.group(1).strip() if m else None


def _markdown_model_title(line: str) -> str | None:
    m = re.match(r"^\s*#\s+(.+?)\s*$", line or "")
    return m.group(1).strip() if m else None


def parse_markdown_import(text: str):
    """Parse Markdown batch import into ModelWrite rows and spec preview rows.

    Supported format per model:
    # NF5280M7
    品牌代码：inspur
    品牌名称：浪潮
    产品类型：服务器
    系列：NF5280
    型号：NF5280M7
    标题：浪潮 NF5280M7

    ## 处理器
    - 处理器：Intel Xeon
    - CPU路数：2
    """
    errors: list[str] = []
    rows: list[ModelWrite] = []
    sheet_rows: list[ImportPreviewSpecRow] = []
    raw = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    if not raw.strip():
        return [], ["Markdown 内容为空"], []

    blocks: list[tuple[int, str, list[str]]] = []
    current_title = ""
    current_start = 1
    current_lines: list[str] = []
    for line_no, line in enumerate(raw.split("\n"), start=1):
        title = _markdown_model_title(line)
        if title:
            if current_title or any(x.strip() for x in current_lines):
                blocks.append((current_start, current_title, current_lines))
            current_title = title
            current_start = line_no
            current_lines = []
        else:
            current_lines.append(line)
    if current_title or any(x.strip() for x in current_lines):
        blocks.append((current_start, current_title, current_lines))
    if not blocks:
        return [], ["未找到 Markdown 型号块；请用一级标题 # 型号 开始每个型号"], []

    for block_index, (start_line, title_line, lines) in enumerate(blocks, start=1):
        meta: dict[str, str] = {}
        specs: list[tuple[int, str, str, str]] = []
        current_group = "其他"
        for offset, line in enumerate(lines, start=1):
            line_no = start_line + offset
            stripped = line.strip()
            if not stripped:
                continue
            section = _markdown_section_title(stripped)
            if section:
                current_group = section
                continue
            kv = _split_markdown_key_value(stripped)
            if not kv:
                continue
            key, value = kv
            mapped = MODEL_META_ALIASES.get(key.strip(), MODEL_META_ALIASES.get(key.strip().lower()))
            if mapped and current_group in {"其他", "基础信息", "基本信息", "型号主数据"}:
                meta[mapped] = value
            else:
                specs.append((line_no, current_group, key, value))
        if title_line and not meta.get("model_name"):
            meta["model_name"] = title_line
        if title_line and not meta.get("title"):
            meta["title"] = title_line
        if not meta.get("source_ref"):
            meta["source_ref"] = "markdown"

        missing = [label for label, field in [("品牌代码", "brand_code"), ("产品类型", "product_type"), ("系列", "series"), ("型号", "model_name")] if not meta.get(field)]
        if missing:
            errors.append(f"Markdown 第 {start_line} 行型号块缺少必要字段：{', '.join(missing)}")
            continue
        model_name = meta["model_name"]
        brand_code, brand_name = normalize_import_brand(meta["brand_code"], meta.get("brand_name"))
        rows.append(ModelWrite(
            brand_code=brand_code,
            brand_name=brand_name,
            product_type=meta["product_type"],
            series=meta["series"],
            model_name=model_name,
            title=meta.get("title") or model_name,
            platform_vendor=meta.get("platform_vendor"),
            generation=meta.get("generation"),
            source_ref=meta.get("source_ref") or "markdown",
            raw_source_id=meta.get("raw_source_id"),
            specifications=[],
        ))
        for line_no, group_name, raw_label, value in specs:
            group_code, field_key, mapped_label = map_field(raw_label)
            effective_group = group_name if group_name not in {"其他", "规格", "参数", "规格参数"} else GROUP_NAME_BY_CODE.get(group_code, "其他")
            sheet_rows.append(ImportPreviewSpecRow(
                row_number=line_no,
                model_name=model_name,
                field_group=effective_group,
                field_label=mapped_label or raw_label,
                field_key=field_key,
                value=value,
                source_ref=meta.get("source_ref") or "markdown",
            ))
    return rows, errors, sheet_rows

def text_of(node) -> str:
    if not node:
        return ""
    return " ".join(node.get_text(" ", strip=True).split())


def clean_value(value: str) -> str:
    value = " ".join(value.split())
    return MISSING if value in {"", "本地资料未提供"} else value


def series_from_model(name: str) -> str:
    match = re.match(r"([A-Za-z]+[0-9]+)", name)
    return match.group(1).upper() if match else name.split()[0]


LENOVO_TS_GLOBAL_MODELS = ("P3 GEN2", "P5 GEN2", "P920", "P620", "PGX", "PX", "P3", "P4", "P5", "P7", "P8")
LENOVO_XINCHUANG_MODELS = ("P3H G1T", "P3H G2T", "P5H G1T", "P7H G1T", "P9Z G1T")


def normalize_lenovo_model_text(value: str) -> str:
    value = value.replace("【企业购】", " ").upper()
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9]+", " ", value)).strip()


def matches_lenovo_model(text: str, names: tuple[str, ...]) -> bool:
    for name in names:
        pattern = re.escape(name).replace(r"\ ", r"\s+")
        if re.search(rf"(^|\s){pattern}($|\s)", text):
            return True
    return False


def lenovo_series_from_model(name: str, product_type: str, category_title: str) -> str:
    normalized = name.replace("【企业购】", "").strip()
    upper = normalized.upper()
    if product_type == "服务器":
        if upper.startswith("WR"):
            return "WR 系列服务器"
        if upper.startswith("SR"):
            return "SR 系列服务器"
        return "其他服务器"
    if "THINKPAD" in upper:
        return "ThinkPad 移动工作站"
    parsed = normalize_lenovo_model_text(normalized)
    if matches_lenovo_model(parsed, LENOVO_XINCHUANG_MODELS):
        return "信创机型"
    if matches_lenovo_model(parsed, LENOVO_TS_GLOBAL_MODELS):
        return "TS全球系列机型"
    return "其他工作站"


def parse_card(card, brand_code: str, brand_name: str, source_ref: str) -> ModelWrite:
    name = text_of(card.select_one(".product-name")) or card.get("data-model") or card.get("id", "")
    title = text_of(card.select_one(".product-title")) or name
    category = card.get("data-category") or text_of(card.find_previous(class_="category-title")) or brand_name
    is_inspur_storage = brand_code == "inspur" and (
        "存储" in category
        or "存储" in title
        or re.match(r"^(HF|AS|NS|JD|INFINISTOR)", name.upper())
    )
    product_type = "工作站" if "工作站" in category else ("存储" if ("存储" in category or is_inspur_storage) else "服务器")
    series_name = card.get("data-family") or series_from_model(name)
    if brand_code == "lenovo":
        prev_section = card.find_parent(class_="category-section")
        category_title = text_of(prev_section.select_one(".category-title")) if prev_section else category
        series_name = lenovo_series_from_model(name, product_type, category_title)
    specs: list[SpecInput] = []
    spec_table = card.select_one(".product-body > table")
    if not spec_table:
        spec_table = card.select_one(".product-body table")
    body = spec_table.find("tbody", recursive=False) if spec_table else None
    rows = body.find_all("tr", recursive=False) if body else (spec_table.find_all("tr", recursive=False) if spec_table else [])
    for row in rows:
        cells = row.find_all(["td", "th"], recursive=False)
        if len(cells) < 2:
            continue
        raw_label = text_of(cells[0])
        raw_value = clean_value(text_of(cells[1]))
        if not raw_label:
            continue
        group_code, field_key, label = map_field(raw_label)
        group_name = {
            "basic": "基础信息",
            "processor": "处理器",
            "memory": "内存",
            "storage": "存储",
            "raid": "RAID",
            "network": "网络",
            "pcie_expansion": "PCIe与扩展",
            "gpu": "GPU",
            "power": "电源",
            "management": "管理",
            "dimension_environment": "尺寸与环境",
            "os_certification": "操作系统与认证",
            "other": "其他",
        }.get(group_code, "其他")
        specs.append(SpecInput(field_key=field_key, label=label, group=group_name, value=raw_value, raw_label=raw_label, source_ref=source_ref))
    if not specs:
        specs.append(SpecInput(field_key="notes", label="备注", group="其他", value=MISSING, raw_label="备注", source_ref=source_ref))
    return ModelWrite(
        brand_code=brand_code,
        brand_name=brand_name,
        product_type=product_type,
        series=series_name,
        model_name=name,
        title=title,
        platform_vendor=card.get("data-platform-vendor"),
        generation=card.get("data-generation"),
        source_ref=source_ref,
        raw_source_id=card.get("id"),
        specifications=specs,
    )


def import_cpu_compatibility(input_dir: Path) -> dict:
    path = input_dir / "inspur-g7-cpu-compatibility.csv"
    if not path.exists():
        return {"available": False, "rows": 0, "fields": []}
    with SessionLocal() as db, path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = 0
        for item in reader:
            existing = db.scalar(
                select(CpuCompatibility).where(
                    CpuCompatibility.server_model == item["server_model"],
                    CpuCompatibility.config_id == item["config_id"],
                    CpuCompatibility.cpu_option_id == item["cpu_option_id"],
                )
            )
            payload = {
                "server_model": item["server_model"],
                "server_id": item["server_id"],
                "config_code": item["config_code"],
                "config_id": item["config_id"],
                "product_name": item["product_name"],
                "cpu_option_id": item["cpu_option_id"],
                "cpu_option_raw": item["cpu_option_raw"],
                "cpu_display": item["cpu_display"],
                "cpu_spec": item["cpu_spec"],
                "source_url": item["source_url"],
                "collected_date": item["collected_date"],
            }
            if existing:
                for key, value in payload.items():
                    setattr(existing, key, value)
            else:
                db.add(CpuCompatibility(**payload))
            rows += 1
        db.commit()
    return {"available": True, "rows": rows, "fields": reader.fieldnames or []}


def import_html(input_dir: Path) -> dict:
    sources = [
        ("inspur", "浪潮", "inspur-index.candidate.html"),
        ("lenovo", "联想", "lenovo-products.fixed.html"),
    ]
    report = {"brands": {}, "field_mappings": {}, "unmapped_fields": [], "missing_parameters": []}
    with SessionLocal() as db:
        seed_spec_template(db)
        ensure_default_api_client(db)
        for brand_code, brand_name, filename in sources:
            path = input_dir / filename
            soup = BeautifulSoup(path.read_text(encoding="utf-8"), "html.parser")
            models = [parse_card(card, brand_code, brand_name, filename) for card in soup.select(".product-card")]
            for model in models:
                upsert_model(db, model, None, "import")
                for spec in model.specifications:
                    report["field_mappings"][spec.raw_label or spec.label] = spec.field_key
                    if spec.field_key.startswith("raw_"):
                        report["unmapped_fields"].append(spec.raw_label or spec.label)
                    if spec.value == MISSING:
                        report["missing_parameters"].append({"brand": brand_name, "model": model.model_name, "field": spec.raw_label or spec.label})
            series_count = len({model.series for model in models})
            report["brands"][brand_code] = {"name": brand_name, "series_count": series_count, "model_count": len(models)}
    report["unmapped_fields"] = sorted(set(report["unmapped_fields"]))
    return report


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", default="/input")
    parser.add_argument("--report", default="/app/import-report.json")
    args = parser.parse_args()
    input_dir = Path(args.input_dir)
    report = import_html(input_dir)
    report["cpu_compatibility"] = import_cpu_compatibility(input_dir)
    Path(args.report).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False))


if __name__ == "__main__":
    main()
