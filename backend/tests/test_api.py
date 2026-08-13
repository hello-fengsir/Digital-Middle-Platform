import os
import tempfile
import time
from http.cookies import SimpleCookie

os.environ["DATABASE_URL"] = f"sqlite:///{tempfile.NamedTemporaryFile(suffix='.db').name}"
os.environ["API_KEY"] = "test-key"
os.environ["ADMIN_USERNAME"] = "admin"
os.environ["ADMIN_PASSWORD_HASH"] = "sha256:" + __import__("hashlib").sha256(b"test-admin-password").hexdigest()
os.environ["ADMIN_SESSION_SECRET"] = "test-session-secret"

from fastapi import HTTPException
from fastapi.testclient import TestClient

from app.catalog import seed_spec_template, upsert_model
from app.db import Base, SessionLocal, engine
from app.main import app
from app.models import Brand, CpuCompatibility, ModelSpecValue, ProductType, Series, SpecDefinition
from app.schemas import ModelWrite, SpecInput
from sqlalchemy import func, select
from app.security import create_admin_session_token, ensure_default_api_client, _b64url_encode
from app.config import settings


def setup_module() -> None:
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    with SessionLocal() as db:
        seed_spec_template(db)
        ensure_default_api_client(db)
        for code, name in [("generic", "示例品牌"), ("inspur", "浪潮"), ("lenovo", "联想"), ("dell", "戴尔")]:
            brand = db.scalar(select(Brand).where(Brand.code == code))
            if not brand:
                db.add(Brand(code=code, name=name, source_name=name))
        db.flush()
        upsert_model(
            db,
            ModelWrite(
                brand_code="inspur",
                brand_name="浪潮",
                product_type="服务器",
                series="NF5280",
                model_name="NF5280M7",
                title="NF5280M7",
                source_ref="test",
                specifications=[
                    SpecInput(field_key="cpu_family", label="处理器", group="处理器", value="Intel Xeon，支持双路 CPU / 双插槽", source_ref="test"),
                    SpecInput(field_key="memory", label="内存", group="内存", value="待补充", source_ref="test"),
                    SpecInput(field_key="pcie_slots", label="PCIe扩展", group="扩展", value="支持 GPU 扩展，4 个 PCIe 扩展槽", source_ref="test"),
                ],
            ),
        )
        upsert_model(
            db,
            ModelWrite(
                brand_code="lenovo",
                brand_name="联想",
                product_type="工作站",
                series="ThinkStation",
                model_name="ThinkStation-PX",
                title="ThinkStation PX 工作站",
                source_ref="test",
                specifications=[SpecInput(field_key="pcie_slots", label="PCIe扩展", group="扩展", value="支持 GPU 扩展，4 个 PCIe 扩展槽", source_ref="test")],
            ),
        )
        server = db.scalar(select(ProductType).where(ProductType.name == "服务器"))
        lenovo = db.scalar(select(Brand).where(Brand.code == "lenovo"))
        if not db.scalar(select(Series).where(Series.brand_id == lenovo.id, Series.product_type_id == server.id, Series.name == "WR5220")):
            db.add(Series(brand_id=lenovo.id, product_type_id=server.id, name="WR5220"))
        db.commit()
        db.add(
            CpuCompatibility(
                server_model="NF5280M7",
                server_id="test-server",
                config_code="NF5280M7_AMD_9124",
                config_id="test-config",
                product_name="SNF5280M7001",
                cpu_option_id="test-cpu",
                cpu_option_raw="AMD EPYC 9124",
                cpu_display="AMD EPYC 9124",
                cpu_spec="3.0GHz / 16C / 64M / 200W",
                source_url="test",
                collected_date="2026-07-22",
            )
        )
        db.commit()


client = TestClient(app)


def auth() -> dict[str, str]:
    return {"X-API-Key": "test-key"}

def admin_token() -> str:
    res = client.post("/api/v1/admin/auth/login", json={"username":"admin","password":"test-admin-password"})
    assert res.status_code == 200
    return res.json()["token"]

def write_auth() -> dict[str, str]:
    return {**auth(), "Authorization": f"Bearer {admin_token()}"}


def test_public_queries() -> None:
    assert client.get("/api/v1/health").status_code == 200
    brands = client.get("/api/v1/brands").json()
    assert [item["name"] for item in brands[:4]] == ["示例品牌", "浪潮", "联想", "戴尔"]
    assert next(item for item in brands if item["code"] == "generic")["model_count"] == 0
    assert any(item["name"] == "服务器" for item in client.get("/api/v1/product-types").json())
    definitions = client.get("/api/v1/spec-definitions").json()
    assert any(item["field_key"] == "cpu_family" and item["group_code"] == "processor" and item["group_name"] == "处理器" for item in definitions)
    link_keys = [item["field_key"] for item in definitions if item["field_key"] in {"selection_notes", "official_params_url", "product_brochure_url", "whitepaper_url"}]
    assert link_keys == ["selection_notes", "official_params_url", "product_brochure_url", "whitepaper_url"]
    series_rows = client.get("/api/v1/series?brand=inspur&product_type=服务器").json()
    assert series_rows[0]["name"] == "NF5280"
    assert series_rows[0]["brand_code"] == "inspur"
    assert series_rows[0]["product_type"] == "服务器"
    models = client.get("/api/v1/models?brand=inspur&keyword=NF5280").json()
    assert models[0]["model_name"] == "NF5280M7"
    detail = client.get(f"/api/v1/models/{models[0]['id']}").json()
    assert detail["specifications"][0]["field_key"] == "cpu_family"
    assert client.get(f"/api/v1/models/{models[0]['id']}/specifications").status_code == 200
    assert client.get("/api/v1/search?q=Xeon").json()[0]["model_name"] == "NF5280M7"
    assert client.get("/api/v1/search?q=9124").json()[0]["model_name"] == "NF5280M7"
    assert client.get("/api/v1/search?q=AMD%20EPYC%209124").json()[0]["model_name"] == "NF5280M7"


def test_public_exact_model_search_ranks_exact_first_and_excludes_negative_notes() -> None:
    with SessionLocal() as db:
        upsert_model(db, ModelWrite(
            brand_code="inspur", brand_name="浪潮", product_type="服务器", series="NF5280",
            model_name="DECOY-M7", title="反向说明样本", source_ref="test",
            specifications=[SpecInput(
                field_key="selection_notes", label="选型说明", group="其他",
                value="低级说明：相近型号，不引用 NF5280M7，不可作为本型号清单", source_ref="test",
            )],
        ))
    for path in ("/api/v1/models?brand=inspur&keyword=NF5280M7", "/api/v1/search?q=NF5280M7&brand=inspur"):
        rows = client.get(path).json()
        assert [item["model_name"] for item in rows] == ["NF5280M7"]


def test_model_detail_filters_blank_spec_values() -> None:
    model = client.get("/api/v1/models?brand=inspur&keyword=NF5280M7").json()[0]
    with SessionLocal() as db:
        definition = db.scalar(select(SpecDefinition).where(SpecDefinition.field_key == "memory"))
        row = db.scalar(select(ModelSpecValue).where(ModelSpecValue.model_id == model["id"], ModelSpecValue.spec_definition_id == definition.id))
        row.value = "  "
        row.raw_value = ""
        db.commit()
    specs = client.get(f"/api/v1/models/{model['id']}/specifications").json()
    assert all(item["value"].strip() for item in specs)


def test_api_key_negative() -> None:
    payload = {"brand_code": "lenovo", "product_type": "服务器", "series": "WR5220", "model_name": "WR1"}
    assert client.post("/api/v1/admin/models", json=payload).status_code == 401
    assert client.post("/api/v1/admin/models", json=payload, headers=auth()).status_code == 401


def test_delete_ai_provider_api_key_requires_admin_bearer() -> None:
    assert client.delete("/api/v1/admin/ai-config/api-key").status_code == 401


def test_delete_ai_provider_api_key_is_atomic_idempotent_and_secret_free(monkeypatch) -> None:
    from app.ai_service import _effective_ai, _xor_secret
    from app.models import AiProviderConfig, AuditLog

    monkeypatch.setattr(settings, "ai_api_key", "environment-key-must-not-return")
    with SessionLocal() as db:
        cfg = db.scalar(select(AiProviderConfig).where(AiProviderConfig.name == "default"))
        if cfg is None:
            cfg = AiProviderConfig(name="default")
            db.add(cfg)
        cfg.base_url = "https://example.com"
        cfg.api_key_cipher = _xor_secret("stored-key-must-not-leak")
        cfg.model = "compatible-model"
        cfg.temperature = "0.35"
        cfg.max_tokens = 777
        cfg.enabled = True
        db.commit()

    headers = {"Authorization": f"Bearer {admin_token()}"}
    first = client.delete("/api/v1/admin/ai-config/api-key", headers=headers)
    assert first.status_code == 200
    assert first.json() == {
        "base_url": "https://example.com", "model": "compatible-model",
        "temperature": 0.35, "max_tokens": 777, "enabled": False,
        "has_api_key": False,
    }
    assert "stored-key" not in first.text
    assert "environment-key" not in first.text
    second = client.delete("/api/v1/admin/ai-config/api-key", headers=headers)
    assert second.status_code == 200
    assert second.json() == first.json()

    with SessionLocal() as db:
        cfg = db.scalar(select(AiProviderConfig).where(AiProviderConfig.name == "default"))
        assert cfg.api_key_cipher == ""
        assert cfg.enabled is False
        assert (cfg.base_url, cfg.model, cfg.temperature, cfg.max_tokens) == (
            "https://example.com", "compatible-model", "0.35", 777,
        )
        assert _effective_ai(db)[1] == ""
        audits = db.scalars(select(AuditLog).where(
            AuditLog.action == "delete_ai_provider_api_key"
        ).order_by(AuditLog.id.desc()).limit(2)).all()
        assert [row.payload["had_api_key"] for row in audits] == [False, True]
        for row in audits:
            assert row.payload["actor"] == "admin"
            serialized = __import__("json").dumps(row.payload).lower()
            assert "stored-key" not in serialized
            assert "environment-key" not in serialized
            assert "cipher" not in serialized


def test_delete_ai_provider_api_key_commit_failure_rolls_back(monkeypatch) -> None:
    from app.ai_service import _xor_secret, delete_ai_provider_api_key
    from app.models import AiProviderConfig

    with SessionLocal() as db:
        cfg = db.scalar(select(AiProviderConfig).where(AiProviderConfig.name == "default"))
        cfg.api_key_cipher = _xor_secret("rollback-key")
        cfg.enabled = True
        db.commit()
        cipher_before = cfg.api_key_cipher
        real_commit = db.commit
        monkeypatch.setattr(db, "commit", lambda: (_ for _ in ()).throw(RuntimeError("forced commit failure")))
        with __import__("pytest").raises(RuntimeError, match="forced commit failure"):
            delete_ai_provider_api_key(db, "admin")
        monkeypatch.setattr(db, "commit", real_commit)
        db.expire_all()
        cfg = db.scalar(select(AiProviderConfig).where(AiProviderConfig.name == "default"))
        assert cfg.api_key_cipher == cipher_before
        assert cfg.enabled is True


def test_existing_empty_ai_config_never_falls_back_to_environment_and_test_fails(monkeypatch) -> None:
    from app.ai_service import _effective_ai, get_ai_config_public, test_ai_config
    from app.models import AiProviderConfig
    from app.schemas import AiConfigTestIn

    monkeypatch.setattr(settings, "ai_api_key", "environment-key-must-stay-unused")
    monkeypatch.setenv("HPL_AI_API_KEY", "alternate-environment-key-must-stay-unused")
    with SessionLocal() as db:
        cfg = db.scalar(select(AiProviderConfig).where(AiProviderConfig.name == "default"))
        cfg.api_key_cipher = ""
        cfg.enabled = False
        db.commit()
        assert get_ai_config_public(db)["has_api_key"] is False
        assert _effective_ai(db)[1] == ""
        with __import__("pytest").raises(HTTPException) as exc:
            test_ai_config(db, AiConfigTestIn())
        assert exc.value.status_code == 400
        assert exc.value.detail == "AI配置不完整"


def test_missing_ai_config_is_seeded_once_from_environment(monkeypatch) -> None:
    from app.ai_service import _get_config, _xor_unsecret
    from app.models import AiProviderConfig

    monkeypatch.setattr(settings, "ai_base_url", "https://example.com")
    monkeypatch.setattr(settings, "ai_api_key", "first-seed-key")
    monkeypatch.setattr(settings, "ai_model", "seed-model")
    with SessionLocal() as db:
        db.query(AiProviderConfig).filter(AiProviderConfig.name == "default").delete()
        db.commit()
        cfg = _get_config(db)
        assert _xor_unsecret(cfg.api_key_cipher) == "first-seed-key"
        cfg_id = cfg.id
        cfg.api_key_cipher = ""
        cfg.enabled = False
        db.commit()
        monkeypatch.setattr(settings, "ai_api_key", "later-environment-key")
        same = _get_config(db)
        assert same.id == cfg_id
        assert same.api_key_cipher == ""


def test_login_only_write_allowed() -> None:
    payload = {"brand_code": "lenovo", "product_type": "服务器", "series": "WR5220", "model_name": "WR_LOGIN_ONLY"}
    created = client.post("/api/v1/admin/models", json=payload, headers={"Authorization": f"Bearer {admin_token()}"})
    assert created.status_code == 200
    model_id = created.json()["id"]
    assert client.delete(f"/api/v1/admin/models/{model_id}", headers={"Authorization": f"Bearer {admin_token()}"}).status_code == 204


def test_admin_auth_endpoints() -> None:
    assert client.post("/api/v1/admin/auth/login", json={"username":"admin","password":"wrong"}).status_code == 401
    login = client.post("/api/v1/admin/auth/login", json={"username":"admin","password":"test-admin-password"})
    assert login.status_code == 200
    token = login.json()["token"]

    # Existing Bearer contract remains valid.
    assert client.get("/api/v1/admin/auth/me", headers={"Authorization": f"Bearer {token}"}).status_code == 200

    set_cookie = login.headers["set-cookie"]
    parsed_cookie = SimpleCookie()
    parsed_cookie.load(set_cookie)
    morsel = parsed_cookie["producthub_admin_session"]
    assert morsel.value == token
    assert morsel["httponly"] is True
    assert morsel["secure"] is True
    assert morsel["samesite"].lower() == "lax"
    assert morsel["path"] == "/"
    assert int(morsel["max-age"]) == settings.admin_session_ttl_seconds

    assert client.get("/api/v1/admin/auth/nginx", cookies={"producthub_admin_session": token}).status_code == 204
    assert client.get("/api/v1/admin/auth/nginx").status_code == 401
    assert client.get("/api/v1/admin/auth/nginx", cookies={"producthub_admin_session": token + "forged"}).status_code == 401

    expired_payload = _b64url_encode(
        __import__("json").dumps({"sub": "admin", "exp": int(time.time()) - 1}, separators=(",", ":"), sort_keys=True).encode()
    )
    signature = __import__("hmac").new(
        settings.admin_session_secret.encode(), expired_payload.encode(), __import__("hashlib").sha256
    ).digest()
    expired_token = f"{expired_payload}.{_b64url_encode(signature)}"
    assert client.get("/api/v1/admin/auth/nginx", cookies={"producthub_admin_session": expired_token}).status_code == 401

    logout = client.post("/api/v1/admin/auth/logout")
    assert logout.status_code == 200
    assert logout.json() == {"ok": True}
    cleared = logout.headers["set-cookie"].lower()
    assert "producthub_admin_session=" in cleared
    assert "max-age=0" in cleared
    assert "httponly" in cleared
    assert "secure" in cleared
    assert "samesite=lax" in cleared
    assert "path=/" in cleared

def test_create_patch_delete_and_specs() -> None:
    payload = {
        "brand_code": "lenovo",
        "brand_name": "联想",
        "product_type": "服务器",
        "series": "WR5220",
        "model_name": "WR5220 G5",
        "title": "WR5220 G5 服务器平台",
        "specifications": [{"field_key": "cpu_family", "label": "处理器", "group": "处理器", "value": "Xeon 6"}],
    }
    created = client.post("/api/v1/admin/models", json=payload, headers=write_auth())
    assert created.status_code == 200
    model_id = created.json()["id"]
    patched = client.patch(
        f"/api/v1/admin/models/{model_id}",
        json={
            "brand_code": "lenovo",
            "brand_name": "联想",
            "product_type": "服务器",
            "series": "WR5220",
            "model_name": "WR5220 G5 Pro",
            "title": "patched",
            "source_ref": "patched-source",
            "raw_source_id": "patched-raw",
        },
        headers=write_auth(),
    )
    assert patched.json()["model_name"] == "WR5220 G5 Pro"
    assert patched.json()["brand_name"] in {"联想", "Lenovo"}
    assert patched.json()["product_type"] == "服务器"
    assert patched.json()["series"] == "WR5220"
    assert patched.json()["title"] == "patched"
    assert patched.json()["source_ref"] == "patched-source"
    assert patched.json()["raw_source_id"] == "patched-raw"
    specs = client.put(
        f"/api/v1/admin/models/{model_id}/specifications",
        json=[{"field_key": "memory", "label": "内存", "group": "内存", "value": "32GB"}],
        headers=write_auth(),
    )
    assert specs.status_code == 200
    assert specs.json()[0]["field_key"] == "memory"
    deleted = client.delete(f"/api/v1/admin/models/{model_id}", headers=write_auth())
    assert deleted.status_code == 204
    assert client.get(f"/api/v1/models/{model_id}").status_code == 404


def test_upsert() -> None:
    payload = {
        "brand_code": "inspur",
        "brand_name": "浪潮",
        "product_type": "服务器",
        "series": "NF5280",
        "model_name": "NF5280M7",
        "title": "upserted",
        "specifications": [{"field_key": "network_interfaces", "label": "网络接口", "group": "网络", "value": "OCP 3.0"}],
    }
    res = client.post("/api/v1/admin/models/upsert", json=payload, headers=write_auth())
    assert res.status_code == 200
    assert res.json()["title"] == "upserted"



def test_excel_import_template_preview_and_run() -> None:
    from io import BytesIO
    from openpyxl import load_workbook
    from app.models import SpecDefinition
    from sqlalchemy import select, func

    headers = {"Authorization": f"Bearer {admin_token()}"}
    template = client.get("/api/v1/admin/import/template", headers=headers)
    assert template.status_code == 200
    wb = load_workbook(BytesIO(template.content), data_only=True)
    assert wb.sheetnames == ["型号主数据", "规格明细"]
    assert [cell.value for cell in wb["型号主数据"][1]] == ["品牌代码", "品牌名称", "产品类型", "系列", "型号", "标题", "平台厂商", "代际", "官网参数链接", "产品彩页链接", "产品技术白皮书下载", "来源引用", "原始来源ID"]
    assert [cell.value for cell in wb["规格明细"][1]] == ["型号", "字段分组", "字段标签", "field_key", "值", "来源引用"]

    files = {"file": ("template.xlsx", template.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    preview = client.post("/api/v1/admin/import/preview", headers=headers, files=files)
    assert preview.status_code == 200
    body = preview.json()
    assert body["total_rows"] == 1
    assert body["valid_rows"] == 1
    assert body["invalid_rows"] == 0
    assert len(body["sheet_rows"]) == 4

    run = client.post("/api/v1/admin/import/run", headers=headers, files={"file": ("template.xlsx", template.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert run.status_code == 200
    assert run.json()["sheet_rows"] == 4
    detail = client.get("/api/v1/models?brand=inspur&keyword=NF5280M7").json()[0]
    specs = client.get(f"/api/v1/models/{detail['id']}/specifications").json()
    values = {item["field_key"]: item["value"] for item in specs}
    assert values["selection_notes"] == "可结合项目场景填写。"
    assert values["official_params_url"] == "https://example.com/params"
    assert values["product_brochure_url"] == "https://example.com/brochure.pdf"
    assert values["whitepaper_url"] == "https://example.com/whitepaper.pdf"

    with SessionLocal() as db:
        assert db.scalar(select(func.count(SpecDefinition.id)).where(SpecDefinition.field_key == "selection_notes")) == 1
        assert db.scalar(select(func.count(SpecDefinition.id)).where(SpecDefinition.field_key == "official_params_url")) == 1
        assert db.scalar(select(func.count(SpecDefinition.id)).where(SpecDefinition.field_key == "product_brochure_url")) == 1
        assert db.scalar(select(func.count(SpecDefinition.id)).where(SpecDefinition.field_key == "whitepaper_url")) == 1


def test_admin_spec_dictionary_is_runtime_locked() -> None:
    headers = {"Authorization": f"Bearer {admin_token()}"}
    definitions = client.get("/api/v1/admin/spec-definitions", headers=headers)
    assert definitions.status_code == 200
    cpu = next(item for item in definitions.json() if item["field_key"] == "cpu_family")
    patched = client.patch(
        f"/api/v1/admin/spec-definitions/{cpu['id']}",
        json={"group_code": "processor", "sort_order": 11, "label": "处理器测试排序"},
        headers=headers,
    )
    assert patched.status_code == 423
    assert patched.json()["detail"]["code"] == "FIELD_DICTIONARY_LOCKED"


def test_new_spec_definition_is_rejected() -> None:
    payload = {
        "brand_code": "inspur", "product_type": "服务器", "series": "NF5280",
        "model_name": "SORT-NEW-FIELD", "title": "SORT-NEW-FIELD",
        "specifications": [{"field_key": "zz_sort_test_field", "label": "排序测试字段", "group": "处理器", "value": "ok"}],
    }
    created = client.post("/api/v1/admin/models", json=payload, headers=write_auth())
    assert created.status_code == 422
    assert created.json()["detail"] == {"code": "INVALID_FIELD_KEY", "field_keys": ["zz_sort_test_field"]}

def test_admin_create_rejects_unknown_brand_and_type_but_allows_new_series() -> None:
    headers = write_auth()
    bad_brand = client.post("/api/v1/admin/models", json={"brand_code": "ghost", "product_type": "服务器", "series": "NF5280", "model_name": "BAD-BRAND"}, headers=headers)
    assert bad_brand.status_code == 400
    bad_type = client.post("/api/v1/admin/models", json={"brand_code": "inspur", "product_type": "不存在类型", "series": "NF5280", "model_name": "BAD-TYPE"}, headers=headers)
    assert bad_type.status_code == 400
    new_series = client.post("/api/v1/admin/models", json={"brand_code": "inspur", "product_type": "服务器", "series": "NEW-CREATE-SERIES", "model_name": "NEW-SERIES-MODEL"}, headers=headers)
    assert new_series.status_code == 200
    assert new_series.json()["series"] == "NEW-CREATE-SERIES"


def test_admin_create_is_immediately_queryable() -> None:
    headers = write_auth()
    payload = {"brand_code": "inspur", "product_type": "服务器", "series": "NF5280", "model_name": "OC-TEMP-QUERY", "title": "OC-TEMP-QUERY"}
    created = client.post("/api/v1/admin/models", json=payload, headers=headers)
    assert created.status_code == 200
    model_id = created.json()["id"]
    listed = client.get("/api/v1/models?brand=inspur&keyword=OC-TEMP-QUERY").json()
    assert any(item["id"] == model_id for item in listed)
    detail = client.get(f"/api/v1/models/{model_id}")
    assert detail.status_code == 200
    assert detail.json()["model_name"] == "OC-TEMP-QUERY"
    assert client.delete(f"/api/v1/admin/models/{model_id}", headers=headers).status_code == 204


def test_admin_patch_existing_model_can_only_switch_to_active_existing_series() -> None:
    headers = write_auth()
    created = client.post("/api/v1/admin/models", json={"brand_code": "lenovo", "product_type": "服务器", "series": "WR5220", "model_name": "PATCH-SERIES-MODEL"}, headers=headers)
    assert created.status_code == 200
    model_id = created.json()["id"]

    bad_series = client.patch(f"/api/v1/admin/models/{model_id}", json={"series": "PATCH-UNKNOWN-SERIES"}, headers=headers)
    assert bad_series.status_code == 400

    with SessionLocal() as db:
        server = db.scalar(select(ProductType).where(ProductType.name == "服务器"))
        lenovo = db.scalar(select(Brand).where(Brand.code == "lenovo"))
        db.add(Series(brand_id=lenovo.id, product_type_id=server.id, name="WR5220-ACTIVE-SWITCH"))
        db.commit()

    switched = client.patch(f"/api/v1/admin/models/{model_id}", json={"series": "WR5220-ACTIVE-SWITCH"}, headers=headers)
    assert switched.status_code == 200
    assert switched.json()["series"] == "WR5220-ACTIVE-SWITCH"


def test_admin_series_delete_empty_204_bound_409() -> None:
    headers = write_auth()
    with SessionLocal() as db:
        server = db.scalar(select(ProductType).where(ProductType.name == "服务器"))
        dell = db.scalar(select(Brand).where(Brand.code == "dell"))
        empty = Series(brand_id=dell.id, product_type_id=server.id, name="EMPTY-DELETE-SERIES")
        db.add(empty)
        db.commit()
        empty_id = empty.id

    assert client.delete(f"/api/v1/admin/series/{empty_id}", headers=headers).status_code == 204

    created = client.post("/api/v1/admin/models", json={"brand_code": "dell", "product_type": "服务器", "series": "BOUND-DELETE-SERIES", "model_name": "BOUND-DELETE-MODEL"}, headers=headers)
    assert created.status_code == 200
    bound_series = next(item for item in client.get("/api/v1/series?brand=dell&product_type=服务器").json() if item["name"] == "BOUND-DELETE-SERIES")
    assert client.delete(f"/api/v1/admin/series/{bound_series['id']}", headers=headers).status_code == 409


def test_normalized_series_name_uniqueness_reuses_existing_series() -> None:
    headers = write_auth()
    first = client.post("/api/v1/admin/models", json={"brand_code": "inspur", "product_type": "服务器", "series": "  NormalizeCaseSeries  ", "model_name": "NORMALIZE-SERIES-1"}, headers=headers)
    assert first.status_code == 200
    assert first.json()["series"] == "NormalizeCaseSeries"
    second = client.post("/api/v1/admin/models", json={"brand_code": "inspur", "product_type": "服务器", "series": "normalizecaseseries", "model_name": "NORMALIZE-SERIES-2"}, headers=headers)
    assert second.status_code == 200
    assert second.json()["series"] == "normalizecaseseries"
    series_rows = [item for item in client.get("/api/v1/series?brand=inspur&product_type=服务器").json() if item["name"].lower() == "normalizecaseseries"]
    assert len(series_rows) == 1
    assert series_rows[0]["model_count"] == 2


def test_public_models_keyword_matches_spec_values_and_definitions() -> None:
    headers = write_auth()
    payload = {
        "brand_code": "inspur",
        "product_type": "服务器",
        "series": "NF5280",
        "model_name": "SPEC-KEYWORD-MODEL",
        "title": "SPEC-KEYWORD-MODEL",
        "specifications": [
            {"field_key": "features", "label": "关键词测试标签", "group": "其他", "value": "value-match-token", "raw_label": "raw-label-token"},
            {"field_key": "notes", "label": "RawValue字段", "group": "其他", "value": "displayed", "raw_label": "RawValue字段"},
        ],
    }
    created = client.post("/api/v1/admin/models", json=payload, headers=headers)
    assert created.status_code == 200
    model_id = created.json()["id"]
    with SessionLocal() as db:
        raw_value_spec = db.scalar(select(ModelSpecValue).join(SpecDefinition).where(ModelSpecValue.model_id == model_id, SpecDefinition.field_key == "notes"))
        raw_value_spec.raw_value = "raw-value-token"
        db.commit()

    for keyword in ["value-match-token", "raw-value-token", "raw-label-token", "特点", "features"]:
        rows = client.get(f"/api/v1/models?brand=inspur&keyword={keyword}").json()
        assert any(item["id"] == model_id for item in rows), keyword



def test_admin_spec_recognition_preview() -> None:
    headers = {"Authorization": f"Bearer {admin_token()}"}
    payload = {
        "brand_code": "inspur",
        "product_type": "服务器",
        "series": "NF5280",
        "model_name": "NF5280M7",
        "raw_text": """处理器：Intel Xeon 6
内存	最高 4TB DDR5
硬盘  支持 24 个 2.5 英寸硬盘
RAID：支持 RAID 0/1/5/10
网卡：4 个千兆网口
电源：双 1600W 白金电源
官网链接：https://example.com/params
未知怪字段：保留但不匹配""",
    }
    res = client.post("/api/v1/admin/spec-recognition/preview", json=payload, headers=headers)
    assert res.status_code == 200
    items = res.json()["items"]
    assert len(items) >= 5
    definitions = {item["field_key"] for item in client.get("/api/v1/spec-definitions").json()}
    keys = {item["matched_field_key"] for item in items}
    required_keys = {"cpu_family", "raid_controller", "network_interfaces", "power_supply", "official_params_url"}
    assert required_keys <= keys
    assert {"memory", "memory_max_capacity"} & keys
    assert {"storage", "drive_bay_25"} & keys
    assert all(item["matched_field_key"] is None or item["matched_field_key"] in definitions for item in items)
    assert any(item["matched_field_key"] is None for item in items)


def test_ai_recommend_dual_socket_has_local_evidence() -> None:
    response = client.post("/api/v1/ai/recommend", json={"message": "双路服务器"})
    assert response.status_code == 200
    body = response.json()
    assert body["match_status"] == "matched"
    assert body["models"]
    assert all(item["product_type"] == "服务器" for item in body["models"])
    assert all(item["evidence"] for item in body["models"])
    assert any("双路" in " ".join(item["evidence"]) for item in body["models"])


def test_ai_recommend_with_evidence_calls_provider(monkeypatch) -> None:
    import app.ai_service as service

    monkeypatch.setattr(service, "_effective_ai", lambda db: ("https://example.com", "key", "model", 0.2, 100, True))
    calls = []
    monkeypatch.setattr(service, "_chat_completion", lambda *args, **kwargs: calls.append(args[3]) or "仅提示核验矩阵风险")
    response = client.post("/api/v1/ai/recommend", json={"message": "双路服务器"})
    assert response.status_code == 200
    body = response.json()
    assert calls and "逐型号本地证据矩阵" in calls[0][0]["content"]
    assert body["source"] == "ai"
    assert body["provenance"] == "ai_used_with_evidence"
    assert "AI 风险提示：仅提示核验矩阵风险" in body["answer"]


def test_ai_recommend_no_match_still_calls_provider_with_bounded_refusal(monkeypatch) -> None:
    import app.ai_service as service

    monkeypatch.setattr(service, "_effective_ai", lambda db: ("https://example.com", "key", "model", 0.2, 100, True))
    calls = []
    monkeypatch.setattr(service, "_chat_completion", lambda *args, **kwargs: calls.append(args[3]) or "越界推荐 ZZZ-9000，内存 100TB")
    response = client.post("/api/v1/ai/recommend", json={"message": "需要量子退火机，100万量子比特，型号ZZZ-NOT-EXIST"})
    assert response.status_code == 200
    body = response.json()
    assert calls and "本地检索无证据" in calls[0][0]["content"]
    assert "需求：" not in calls[0][0]["content"]
    assert body["match_status"] == "no_match"
    assert body["models"] == []
    assert body["source"] == "ai"
    assert body["provenance"] == "ai_used_no_evidence_refusal"
    assert "越界推荐" not in body["answer"]
    assert service.NO_EVIDENCE_REFUSAL in body["answer"]
    assert "未找到" in body["answer"]


def test_ai_recommend_unparsed_calls_provider_without_requirement_details(monkeypatch) -> None:
    import app.ai_service as service

    monkeypatch.setattr(service, "_effective_ai", lambda db: ("https://example.com", "key", "model", 0.2, 100, True))
    calls = []
    monkeypatch.setattr(service, "_chat_completion", lambda *args, **kwargs: calls.append(args[3]) or service.UNPARSED_REFUSAL)
    response = client.post("/api/v1/ai/recommend", json={"message": "浪潮服务器，要求热插拔风扇"})
    assert response.status_code == 200
    body = response.json()
    assert calls and "存在未解析条件" in calls[0][0]["content"]
    assert "浪潮" not in calls[0][0]["content"] and "热插拔" not in calls[0][0]["content"]
    assert body["source"] == "ai"
    assert body["provenance"] == "ai_used_no_evidence_refusal"
    assert service.UNPARSED_REFUSAL in body["answer"]


def test_ai_recommend_provider_timeout_and_error_fall_back_to_local(monkeypatch) -> None:
    import app.ai_service as service

    monkeypatch.setattr(service, "_effective_ai", lambda db: ("https://example.com", "key", "model", 0.2, 100, True))
    for status_code in (504, 502):
        monkeypatch.setattr(service, "_chat_completion", lambda *args, _status=status_code, **kwargs: (_ for _ in ()).throw(HTTPException(status_code=_status, detail="provider failure")))
        response = client.post("/api/v1/ai/recommend", json={"message": "双路服务器"})
        assert response.status_code == 200
        body = response.json()
        assert body["source"] == "local"
        assert body["provenance"] == "ai_provider_failed"
        assert body["warning"] == service.PUBLIC_AI_WARNING


def test_ai_recommend_cross_brand_gpu_expansion_coverage() -> None:
    response = client.post("/api/v1/ai/recommend", json={"message": "比较联想工作站和浪潮服务器，要求GPU扩展"})
    assert response.status_code == 200
    body = response.json()
    assert body["match_status"] == "matched"
    assert {(item["brand_code"], item["product_type"]) for item in body["models"]} == {
        ("lenovo", "工作站"), ("inspur", "服务器"),
    }
    assert set(body["coverage"]["requested_brands"]) == {"lenovo", "inspur"}
    assert set(body["coverage"]["covered_brands"]) == {"lenovo", "inspur"}
    assert body["coverage"]["uncovered_brands"] == []
    assert len(body["coverage"]["brand_results"]) == 2


def test_ai_recommend_message_length_boundary_rejected_before_handler(monkeypatch) -> None:
    import app.routes.public as public_routes

    monkeypatch.setattr(public_routes, "recommend_models", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("handler must not run")), raising=False)
    assert client.post("/api/v1/ai/recommend", json={"message": ""}).status_code == 422
    response = client.post("/api/v1/ai/recommend", json={"message": "字" * 4001})
    assert response.status_code == 422
    assert response.json()["detail"][0]["type"] == "string_too_long"



def test_controlled_lifecycle_enums_and_structured_public_badges() -> None:
    headers = write_auth()
    expected = {
        "npi": "新品", "rts": "在售", "rtq": "可报价",
        "eos": "停止接单", "eol": "停售",
    }
    ids = []
    for lifecycle, label in expected.items():
        created = client.post("/api/v1/admin/models", json={
            "brand_code": "inspur", "product_type": "服务器", "series": "NF5280",
            "model_name": f"LIFECYCLE-{lifecycle.upper()}",
            "lifecycle_status": lifecycle,
            "business_tags": ["featured"] if lifecycle == "rts" else [],
        }, headers=headers)
        assert created.status_code == 200, created.text
        body = created.json()
        ids.append(body["id"])
        assert body["lifecycle_status"] == lifecycle
        assert body["badges"][0] == {"kind": "lifecycle", "code": lifecycle, "label": label}
        if lifecycle == "rts":
            assert body["business_tags"] == ["featured"]
            assert body["badges"][1] == {"kind": "business", "code": "featured", "label": "主推"}
        summary = next(row for row in client.get(f"/api/v1/models?keyword=LIFECYCLE-{lifecycle.upper()}").json() if row["id"] == body["id"])
        detail = client.get(f"/api/v1/models/{body['id']}").json()
        assert summary["badges"] == body["badges"]
        assert detail["badges"] == body["badges"]
    for model_id in ids:
        assert client.delete(f"/api/v1/admin/models/{model_id}", headers=headers).status_code == 204


def test_unknown_lifecycle_or_business_tag_is_422_and_patch_is_atomic() -> None:
    headers = write_auth()
    created = client.post("/api/v1/admin/models", json={
        "brand_code": "inspur", "product_type": "服务器", "series": "NF5280",
        "model_name": "LIFECYCLE-ATOMIC", "title": "before", "lifecycle_status": "rtq",
    }, headers=headers).json()
    model_id = created["id"]
    bad_lifecycle = client.patch(f"/api/v1/admin/models/{model_id}", json={
        "title": "must-not-write", "lifecycle_status": "任意文本",
    }, headers=headers)
    assert bad_lifecycle.status_code == 422
    bad_tag = client.patch(f"/api/v1/admin/models/{model_id}", json={
        "title": "must-not-write-either", "business_tags": ["featured", "guessed"],
    }, headers=headers)
    assert bad_tag.status_code == 422
    unchanged = client.get(f"/api/v1/admin/models/{model_id}", headers=headers).json()
    assert unchanged["title"] == "before"
    assert unchanged["lifecycle_status"] == "rtq"
    assert unchanged["business_tags"] == []
    with SessionLocal() as db:
        from app.models import AuditLog, ModelBusinessTag
        from sqlalchemy import func
        assert db.scalar(select(func.count(ModelBusinessTag.id)).where(ModelBusinessTag.model_id == model_id)) == 0
        assert db.scalar(select(func.count(AuditLog.id)).where(AuditLog.entity_id == model_id, AuditLog.action == "patch")) == 0
    assert client.delete(f"/api/v1/admin/models/{model_id}", headers=headers).status_code == 204


def test_eos_eol_are_searchable_and_detail_visible_but_excluded_from_ai_candidates() -> None:
    from app.ai_service import _candidate_models
    headers = write_auth()
    ids = []
    for lifecycle in ("eos", "eol"):
        body = client.post("/api/v1/admin/models", json={
            "brand_code": "inspur", "product_type": "服务器", "series": "NF5280",
            "model_name": f"AI-EXCLUDE-{lifecycle.upper()}", "lifecycle_status": lifecycle,
        }, headers=headers).json()
        ids.append(body["id"])
        assert any(row["id"] == body["id"] for row in client.get(f"/api/v1/search?q=AI-EXCLUDE-{lifecycle.upper()}").json())
        assert client.get(f"/api/v1/models/{body['id']}").status_code == 200
        with SessionLocal() as db:
            assert all(model.id != body["id"] for model in _candidate_models(db, f"AI-EXCLUDE-{lifecycle.upper()}"))
    for model_id in ids:
        assert client.delete(f"/api/v1/admin/models/{model_id}", headers=headers).status_code == 204


def test_legacy_model_without_evidence_remains_unknown_without_lifecycle_badge() -> None:
    row = client.get("/api/v1/models?brand=inspur&keyword=NF5280M7").json()[0]
    assert row["lifecycle_status"] is None
    assert row["business_tags"] == []
    assert row["badges"] == []



def test_omitted_lifecycle_is_unknown_without_public_lifecycle_badge_and_ai_can_consider_it() -> None:
    from app.ai_service import _candidate_models
    headers = write_auth()
    response = client.post("/api/v1/admin/models", json={
        "brand_code": "inspur", "product_type": "服务器", "series": "NF5280",
        "model_name": "LIFECYCLE-UNKNOWN-NULL", "title": "unknown lifecycle",
    }, headers=headers)
    assert response.status_code == 200, response.text
    body = response.json()
    model_id = body["id"]
    assert body["lifecycle_status"] is None
    assert not any(badge["kind"] == "lifecycle" for badge in body["badges"])
    summary = next(row for row in client.get("/api/v1/models?keyword=LIFECYCLE-UNKNOWN-NULL").json() if row["id"] == model_id)
    assert summary["lifecycle_status"] is None
    assert summary["badges"] == []
    with SessionLocal() as db:
        assert any(model.id == model_id for model in _candidate_models(db, "LIFECYCLE-UNKNOWN-NULL"))
    assert client.delete(f"/api/v1/admin/models/{model_id}", headers=headers).status_code == 204


def test_explicit_null_lifecycle_is_422_with_zero_create_or_patch_writes() -> None:
    headers = write_auth()
    marker = "LIFECYCLE-EXPLICIT-NULL"
    with SessionLocal() as db:
        from app.models import AuditLog, Model
        before_models = db.scalar(select(func.count(Model.id)).where(Model.model_name == marker))
        before_audits = db.scalar(select(func.count(AuditLog.id)))
    bad_create = client.post("/api/v1/admin/models", json={
        "brand_code": "inspur", "product_type": "服务器", "series": "NF5280",
        "model_name": marker, "lifecycle_status": None,
    }, headers=headers)
    assert bad_create.status_code == 422
    with SessionLocal() as db:
        assert db.scalar(select(func.count(Model.id)).where(Model.model_name == marker)) == before_models
        assert db.scalar(select(func.count(AuditLog.id))) == before_audits

    created = client.post("/api/v1/admin/models", json={
        "brand_code": "inspur", "product_type": "服务器", "series": "NF5280",
        "model_name": marker, "title": "before", "lifecycle_status": "rts",
    }, headers=headers).json()
    bad_patch = client.patch(f"/api/v1/admin/models/{created['id']}", json={
        "title": "must-not-write", "lifecycle_status": None,
    }, headers=headers)
    assert bad_patch.status_code == 422
    unchanged = client.get(f"/api/v1/admin/models/{created['id']}", headers=headers).json()
    assert unchanged["title"] == "before"
    assert unchanged["lifecycle_status"] == "rts"
    with SessionLocal() as db:
        from app.models import AuditLog
        assert db.scalar(select(func.count(AuditLog.id)).where(AuditLog.entity_id == created["id"], AuditLog.action == "patch")) == 0
    assert client.delete(f"/api/v1/admin/models/{created['id']}", headers=headers).status_code == 204


def test_explicit_compatible_gpu_save_contract() -> None:
    from app.models import AuditLog, Model, ModelCompatibleGpu

    with SessionLocal() as db:
        gpu1 = upsert_model(db, ModelWrite(brand_code="accessory", brand_name="配件", product_type="显卡", series="GPU", model_name="API-GPU-1", title="GPU 1", source_ref="test"))
        gpu2 = upsert_model(db, ModelWrite(brand_code="accessory", brand_name="配件", product_type="显卡", series="GPU", model_name="API-GPU-2", title="GPU 2", source_ref="test"))
        inactive = upsert_model(db, ModelWrite(brand_code="accessory", brand_name="配件", product_type="显卡", series="GPU", model_name="API-GPU-INACTIVE", title="inactive", source_ref="test"))
        inactive.status = "deleted"
        db.commit()
        server = db.scalar(select(Model).where(Model.model_name == "NF5280M7"))
        ids = server.id, gpu1.id, gpu2.id, inactive.id

    server_id, gpu1_id, gpu2_id, inactive_id = ids
    url = f"/api/v1/admin/models/{server_id}/compatible-gpus"
    assert client.put(url, json={"compatible_gpu_ids": [gpu1_id]}).status_code == 401
    assert client.put(url, json={"compatible_gpu_ids": [gpu1_id], "unexpected": True}, headers=write_auth()).status_code == 422

    saved = client.put(url, json={"compatible_gpu_ids": [gpu2_id, gpu1_id, gpu1_id]}, headers=write_auth())
    assert saved.status_code == 200, saved.text
    assert {item["id"] for item in saved.json()} == {gpu1_id, gpu2_id}
    assert len(saved.json()) == 2
    assert {item["id"] for item in client.get(f"/api/v1/models/{server_id}").json()["compatible_gpus"]} == {gpu1_id, gpu2_id}

    with SessionLocal() as db:
        before = list(db.scalars(select(ModelCompatibleGpu.gpu_model_id).where(ModelCompatibleGpu.model_id == server_id)).all())
        audit = db.scalar(select(AuditLog).where(AuditLog.entity_id == server_id, AuditLog.action == "replace_compatible_gpus").order_by(AuditLog.id.desc()))
        assert audit.payload["compatible_gpu_ids"] == [gpu2_id, gpu1_id]

    for invalid_id in (999999999, inactive_id):
        assert client.patch(url, json={"compatible_gpu_ids": [invalid_id]}, headers=write_auth()).status_code == 400
        with SessionLocal() as db:
            assert list(db.scalars(select(ModelCompatibleGpu.gpu_model_id).where(ModelCompatibleGpu.model_id == server_id)).all()) == before

    assert client.put(url, json={"compatible_gpu_ids": [server_id]}, headers=write_auth()).status_code == 400
    assert client.put(f"/api/v1/admin/models/{gpu1_id}/compatible-gpus", json={"compatible_gpu_ids": [gpu2_id]}, headers=write_auth()).status_code == 400
    basic_patch = client.patch(f"/api/v1/admin/models/{server_id}", json={"title": "base patch remains compatible"}, headers=write_auth())
    assert basic_patch.status_code == 200
    assert {item["id"] for item in basic_patch.json()["compatible_gpus"]} == {gpu1_id, gpu2_id}

    cleared = client.put(url, json={"compatible_gpu_ids": []}, headers=write_auth())
    assert cleared.status_code == 200 and cleared.json() == []
    assert client.get(f"/api/v1/models/{server_id}").json()["compatible_gpus"] == []
